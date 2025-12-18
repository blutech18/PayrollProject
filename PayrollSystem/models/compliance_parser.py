"""
Compliance File Parser for PROLY Payroll System.
Parses uploaded compliance files (CSV, Excel) and extracts rate tables.
"""

from __future__ import annotations

import csv
import os
from typing import Dict, List, Optional, Tuple
import logging

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


logger = logging.getLogger(__name__)


def parse_sss_file(file_path: str) -> Dict:
    """
    Parse SSS contribution table file.
    
    Expected format (CSV):
    Salary Range,Employee Share,Employer Share,Total
    0-3250,135.00,135.00,270.00
    3251-3750,157.50,157.50,315.00
    
    Returns:
        Dictionary with salary brackets and contribution amounts
    """
    brackets = []
    
    try:
        if file_path.endswith('.csv'):
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    # Parse salary range
                    salary_range = row.get('Salary Range', '')
                    if '-' in salary_range:
                        min_sal, max_sal = salary_range.split('-')
                        brackets.append({
                            'min': float(min_sal.strip()),
                            'max': float(max_sal.strip()),
                            'employee_share': float(row.get('Employee Share', 0)),
                            'employer_share': float(row.get('Employer Share', 0))
                        })
        elif file_path.endswith(('.xlsx', '.xls')) and EXCEL_AVAILABLE:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    salary_range = str(row[0])
                    if '-' in salary_range:
                        min_sal, max_sal = salary_range.split('-')
                        brackets.append({
                            'min': float(min_sal.strip()),
                            'max': float(max_sal.strip()),
                            'employee_share': float(row[1] or 0),
                            'employer_share': float(row[2] or 0)
                        })
    except Exception as e:
        logger.exception("Error parsing SSS file: %s", e)
        return {'brackets': []}
    
    return {'brackets': brackets, 'type': 'SSS'}


def parse_philhealth_file(file_path: str) -> Dict:
    """
    Parse PhilHealth rates file.
    
    Expected format (CSV):
    Monthly Salary,Employee Share,Employer Share
    0-10000,300.00,300.00
    10001-80000,3.00%,3.00%
    80001+,2400.00,2400.00
    
    Returns:
        Dictionary with salary brackets and rates
    """
    brackets = []
    
    try:
        if file_path.endswith('.csv'):
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    salary_range = row.get('Monthly Salary', '')
                    employee_share = row.get('Employee Share', '0')
                    employer_share = row.get('Employer Share', '0')
                    
                    # Parse percentage or fixed amount
                    if '%' in employee_share:
                        emp_rate = float(employee_share.replace('%', '')) / 100
                        emp_fixed = None
                    else:
                        emp_rate = None
                        emp_fixed = float(employee_share)
                    
                    if '%' in employer_share:
                        emp_rate_emp = float(employer_share.replace('%', '')) / 100
                        emp_fixed_emp = None
                    else:
                        emp_rate_emp = None
                        emp_fixed_emp = float(employer_share)
                    
                    # Parse range
                    if '-' in salary_range:
                        min_sal, max_sal = salary_range.split('-')
                        brackets.append({
                            'min': float(min_sal.strip()),
                            'max': float(max_sal.strip()) if max_sal.strip() != '+' else float('inf'),
                            'employee_rate': emp_rate,
                            'employee_fixed': emp_fixed,
                            'employer_rate': emp_rate_emp,
                            'employer_fixed': emp_fixed_emp
                        })
        elif file_path.endswith(('.xlsx', '.xls')) and EXCEL_AVAILABLE:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    salary_range = str(row[0])
                    if '-' in salary_range:
                        min_sal, max_sal = salary_range.split('-')
                        brackets.append({
                            'min': float(min_sal.strip()),
                            'max': float(max_sal.strip()) if max_sal.strip() != '+' else float('inf'),
                            'employee_rate': float(row[1] or 0) / 100 if '%' in str(row[1] or '') else None,
                            'employee_fixed': float(row[1] or 0) if '%' not in str(row[1] or '') else None
                        })
    except Exception as e:
        logger.exception("Error parsing PhilHealth file: %s", e)
        return {'brackets': []}
    
    return {'brackets': brackets, 'type': 'PHILHEALTH'}


def parse_pagibig_file(file_path: str) -> Dict:
    """
    Parse Pag-IBIG rates file.
    
    Expected format (CSV):
    Monthly Salary,Employee Share,Employer Share
    0-1500,1.00%,2.00%
    1501+,2.00%,2.00%
    Maximum,100.00,100.00
    
    Returns:
        Dictionary with salary brackets and rates
    """
    brackets = []
    max_contribution = 100.0
    
    try:
        if file_path.endswith('.csv'):
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    salary_range = row.get('Monthly Salary', '')
                    if 'Maximum' in salary_range or 'Max' in salary_range:
                        max_contribution = float(row.get('Employee Share', 100))
                    else:
                        employee_share = row.get('Employee Share', '0')
                        if '%' in employee_share:
                            emp_rate = float(employee_share.replace('%', '')) / 100
                        else:
                            emp_rate = None
                        
                        if '-' in salary_range:
                            min_sal, max_sal = salary_range.split('-')
                            brackets.append({
                                'min': float(min_sal.strip()),
                                'max': float(max_sal.strip()) if max_sal.strip() != '+' else float('inf'),
                                'employee_rate': emp_rate
                            })
        elif file_path.endswith(('.xlsx', '.xls')) and EXCEL_AVAILABLE:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    salary_range = str(row[0])
                    if 'Maximum' in salary_range or 'Max' in salary_range:
                        max_contribution = float(row[1] or 100)
                    elif '-' in salary_range:
                        min_sal, max_sal = salary_range.split('-')
                        emp_share = str(row[1] or '0')
                        emp_rate = float(emp_share.replace('%', '')) / 100 if '%' in emp_share else None
                        brackets.append({
                            'min': float(min_sal.strip()),
                            'max': float(max_sal.strip()) if max_sal.strip() != '+' else float('inf'),
                            'employee_rate': emp_rate
                        })
    except Exception as e:
        logger.exception("Error parsing Pag-IBIG file: %s", e)
        return {'brackets': [], 'max_contribution': 100.0}
    
    return {'brackets': brackets, 'max_contribution': max_contribution, 'type': 'PAGIBIG'}


def parse_tax_file(file_path: str) -> Dict:
    """
    Parse BIR tax table file.
    
    Expected format (CSV):
    Taxable Income From,To,Base Tax,Rate Over
    0,20833,0.00,0%
    20833,33332,0.00,20%
    33332,66666,2500.00,25%
    
    Returns:
        Dictionary with tax brackets
    """
    brackets = []
    
    try:
        if file_path.endswith('.csv'):
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    brackets.append({
                        'from': float(row.get('Taxable Income From', 0)),
                        'to': float(row.get('To', float('inf'))),
                        'base_tax': float(row.get('Base Tax', 0)),
                        'rate': float(row.get('Rate Over', '0%').replace('%', '')) / 100
                    })
        elif file_path.endswith(('.xlsx', '.xls')) and EXCEL_AVAILABLE:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    rate_str = str(row[3] or '0%')
                    brackets.append({
                        'from': float(row[0] or 0),
                        'to': float(row[1] or float('inf')),
                        'base_tax': float(row[2] or 0),
                        'rate': float(rate_str.replace('%', '')) / 100
                    })
    except Exception as e:
        logger.exception("Error parsing Tax file: %s", e)
        return {'brackets': []}
    
    return {'brackets': brackets, 'type': 'TAX'}


def parse_compliance_file(file_type: str, file_path: str) -> Optional[Dict]:
    """
    Parse a compliance file based on its type.
    
    Args:
        file_type: Type of compliance file ('BIR Tax', 'SSS Contributions', etc.)
        file_path: Path to the file
    
    Returns:
        Parsed data dictionary or None if parsing fails
    """
    if not os.path.exists(file_path):
        return None
    
    type_mapping = {
        'SSS Contributions': parse_sss_file,
        'PhilHealth Rates': parse_philhealth_file,
        'Pag-IBIG Rates': parse_pagibig_file,
        'BIR Tax': parse_tax_file
    }
    
    parser = type_mapping.get(file_type)
    if parser:
        return parser(file_path)
    
    return None

